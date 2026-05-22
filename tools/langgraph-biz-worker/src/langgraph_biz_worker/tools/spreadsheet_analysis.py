"""Deterministic spreadsheet attachment analysis tools."""

from __future__ import annotations

import base64
import csv
import datetime as dt
import io
import json
import logging
import re
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit
from urllib.request import Request, urlopen

from ..runtime.attachment_context import build_attachment_evidence

logger = logging.getLogger(__name__)

_SUPPORTED_OPERATIONS = frozenset({"summary", "preview", "read_range", "extract_rows"})
_XLSX_MIME_TYPES = frozenset({
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
})
_CSV_MIME_TYPES = frozenset({
    "text/csv",
    "application/csv",
})
_UNSUPPORTED_EXCEL_SUFFIXES = frozenset({".xls", ".xlsm", ".xlsb"})

_MAX_ATTACHMENT_BYTES = 10 * 1024 * 1024
_MAX_RETURN_ROWS = 100
_MAX_RETURN_COLUMNS = 60
_DEFAULT_PREVIEW_ROWS = 20
_DEFAULT_EXTRACT_ROWS = 50
_MAX_CELL_CHARS = 500
_MAX_RESULT_CHARS = 24_000
_DOWNLOAD_TIMEOUT_SECONDS = 20


def analyze_spreadsheet(
    args: dict[str, Any],
    runtime_context: dict[str, Any] | None,
    *,
    artifact_store: Any | None = None,
    account_id: str | None = None,
    task_id: str | None = None,
) -> dict[str, Any]:
    """Analyze one spreadsheet attachment by id.

    The LLM supplies only ``attachment_id`` and operation arguments. File
    bytes are resolved from runtime attachment metadata so signed URLs or
    local paths do not appear in prompts or tool-call arguments.
    """
    context = runtime_context or {}
    attachment_id = _text(args.get("attachment_id") or args.get("attachmentId") or args.get("id"))
    if not attachment_id:
        return {"ok": False, "error": "ATTACHMENT_ID_REQUIRED: attachment_id is required"}

    operation = _text(args.get("operation") or "summary").lower() or "summary"
    if operation not in _SUPPORTED_OPERATIONS:
        return {
            "ok": False,
            "error": f"UNSUPPORTED_SPREADSHEET_OPERATION: {operation}",
            "supported_operations": sorted(_SUPPORTED_OPERATIONS),
        }

    attachment = _find_attachment(context.get("attachments"), attachment_id)
    if attachment is None:
        return {"ok": False, "error": f"ATTACHMENT_NOT_FOUND: {attachment_id}"}

    spreadsheet_format = _spreadsheet_format(attachment)
    if spreadsheet_format is None:
        suffix = _attachment_suffix(attachment)
        if suffix in _UNSUPPORTED_EXCEL_SUFFIXES:
            return {"ok": False, "error": f"UNSUPPORTED_SPREADSHEET_FORMAT: {suffix}"}
        return {"ok": False, "error": f"UNSUPPORTED_ATTACHMENT_TYPE: {attachment_id}"}

    content_result = _read_attachment_bytes(attachment, attachment_id)
    if not content_result.get("ok"):
        return content_result
    content = content_result["content"]

    try:
        if spreadsheet_format == "xlsx":
            result = _analyze_xlsx(content, attachment, attachment_id, operation, args)
        else:
            result = _analyze_csv(content, attachment, attachment_id, operation, args)
    except SpreadsheetAnalysisError as exc:
        return {"ok": False, "error": f"{exc.code}: {exc.detail}"}
    except Exception as exc:
        logger.warning(
            "Spreadsheet analysis failed: task_id=%s attachment_id=%s operation=%s error=%s",
            _text(context.get("task_id") or context.get("taskId")),
            attachment_id,
            operation,
            _safe_error_summary(exc),
        )
        return {"ok": False, "error": f"SPREADSHEET_PARSE_FAILED: {_safe_error_summary(exc)}"}

    result["attachment_evidence"] = build_attachment_evidence([attachment])
    return _maybe_externalize_result(
        result,
        artifact_store=artifact_store,
        account_id=account_id,
        task_id=task_id,
    )


class SpreadsheetAnalysisError(Exception):
    def __init__(self, code: str, detail: str) -> None:
        self.code = code
        self.detail = detail
        super().__init__(f"{code}: {detail}")


def _analyze_xlsx(
    content: bytes,
    attachment: dict[str, Any],
    attachment_id: str,
    operation: str,
    args: dict[str, Any],
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries
    except ImportError as exc:
        raise SpreadsheetAnalysisError(
            "SPREADSHEET_DEPENDENCY_MISSING",
            "openpyxl is required for .xlsx parsing",
        ) from exc

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise SpreadsheetAnalysisError("SPREADSHEET_OPEN_FAILED", _safe_error_summary(exc)) from exc

    formula_workbook = None
    try:
        try:
            formula_workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
        except Exception:
            formula_workbook = None

        if operation == "summary":
            return _xlsx_summary(workbook, attachment, attachment_id)

        sheet = _select_sheet(workbook, args.get("sheet"))
        formula_sheet = formula_workbook[sheet.title] if formula_workbook and sheet.title in formula_workbook else None

        if operation == "preview":
            limit = _int_arg(args, "limit", _DEFAULT_PREVIEW_ROWS, 1, _MAX_RETURN_ROWS)
            max_col = min(sheet.max_column or 1, _MAX_RETURN_COLUMNS)
            return _xlsx_range_result(
                workbook,
                attachment,
                attachment_id,
                operation,
                sheet.title,
                1,
                1,
                max_col,
                min(sheet.max_row or 1, limit),
                formula_sheet=formula_sheet,
                include_formula_text=_bool_option(args, "include_formulas", False),
            )

        if operation == "read_range":
            range_text = _text(args.get("range") or args.get("cell_range") or args.get("cellRange"))
            if not range_text:
                raise SpreadsheetAnalysisError("RANGE_REQUIRED", "range is required for read_range")
            try:
                min_col, min_row, max_col, max_row = range_boundaries(range_text)
            except ValueError as exc:
                raise SpreadsheetAnalysisError("INVALID_RANGE", range_text) from exc
            return _xlsx_range_result(
                workbook,
                attachment,
                attachment_id,
                operation,
                sheet.title,
                min_col,
                min_row,
                max_col,
                max_row,
                formula_sheet=formula_sheet,
                include_formula_text=_bool_option(args, "include_formulas", False),
            )

        if operation == "extract_rows":
            return _xlsx_extract_rows(
                workbook,
                attachment,
                attachment_id,
                sheet.title,
                args,
                formula_sheet=formula_sheet,
            )

        raise SpreadsheetAnalysisError("UNSUPPORTED_SPREADSHEET_OPERATION", operation)
    finally:
        workbook.close()
        if formula_workbook is not None:
            formula_workbook.close()


def _xlsx_summary(workbook: Any, attachment: dict[str, Any], attachment_id: str) -> dict[str, Any]:
    sheets: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        dimension = ""
        try:
            dimension = worksheet.calculate_dimension()
        except Exception:
            dimension = ""
        sheets.append({
            "name": worksheet.title,
            "state": getattr(worksheet, "sheet_state", "visible"),
            "max_row": int(worksheet.max_row or 0),
            "max_column": int(worksheet.max_column or 0),
            "dimension": dimension,
        })
    return {
        "ok": True,
        "attachment_id": attachment_id,
        "operation": "summary",
        "workbook": {
            "file_name": _attachment_name(attachment),
            "format": "xlsx",
            "sheet_count": len(sheets),
        },
        "sheets": sheets,
        "truncated": False,
        "warnings": [],
    }


def _xlsx_range_result(
    workbook: Any,
    attachment: dict[str, Any],
    attachment_id: str,
    operation: str,
    sheet_name: str,
    min_col: int,
    min_row: int,
    max_col: int,
    max_row: int,
    *,
    formula_sheet: Any | None,
    include_formula_text: bool,
) -> dict[str, Any]:
    from openpyxl.utils.cell import get_column_letter

    worksheet = workbook[sheet_name]
    warnings: list[str] = []
    original_rows = max(0, max_row - min_row + 1)
    original_cols = max(0, max_col - min_col + 1)
    if original_rows > _MAX_RETURN_ROWS:
        max_row = min_row + _MAX_RETURN_ROWS - 1
        warnings.append(f"ROW_LIMIT_APPLIED: returned first {_MAX_RETURN_ROWS} rows from requested range")
    if original_cols > _MAX_RETURN_COLUMNS:
        max_col = min_col + _MAX_RETURN_COLUMNS - 1
        warnings.append(f"COLUMN_LIMIT_APPLIED: returned first {_MAX_RETURN_COLUMNS} columns from requested range")

    formula_cells = _formula_coordinates(formula_sheet, min_row, max_row, min_col, max_col)
    if formula_cells:
        warnings.append("FORMULA_CELLS_PRESENT: formulas were not executed; cached values were returned")

    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        cells = [
            _cell_payload(cell, formula_cells=formula_cells, include_formula_text=include_formula_text)
            for cell in row
        ]
        rows.append({
            "row_index": row[0].row if row else None,
            "values": [cell["value"] for cell in cells],
            "cells": cells,
        })

    columns = [
        {
            "index": col,
            "letter": get_column_letter(col),
        }
        for col in range(min_col, max_col + 1)
    ]
    return {
        "ok": True,
        "attachment_id": attachment_id,
        "operation": operation,
        "workbook": {
            "file_name": _attachment_name(attachment),
            "format": "xlsx",
            "sheet_count": len(workbook.sheetnames),
        },
        "sheet": {
            "name": sheet_name,
            "max_row": int(worksheet.max_row or 0),
            "max_column": int(worksheet.max_column or 0),
        },
        "range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
        "columns": columns,
        "rows": rows,
        "truncated": bool(original_rows > _MAX_RETURN_ROWS or original_cols > _MAX_RETURN_COLUMNS),
        "warnings": warnings,
    }


def _xlsx_extract_rows(
    workbook: Any,
    attachment: dict[str, Any],
    attachment_id: str,
    sheet_name: str,
    args: dict[str, Any],
    *,
    formula_sheet: Any | None,
) -> dict[str, Any]:
    from openpyxl.utils.cell import get_column_letter

    worksheet = workbook[sheet_name]
    header_row = _int_arg(args, "header_row", 1, 1, int(worksheet.max_row or 1))
    offset = _int_arg(args, "offset", 0, 0, 100_000)
    limit = _int_arg(args, "limit", _DEFAULT_EXTRACT_ROWS, 1, _MAX_RETURN_ROWS)
    max_col = min(int(worksheet.max_column or 1), _MAX_RETURN_COLUMNS)
    headers = _xlsx_row_values(worksheet, header_row, 1, max_col)
    normalized_headers = _normalize_headers(headers)
    start_row = header_row + 1 + offset
    end_row = min(start_row + limit - 1, int(worksheet.max_row or start_row))
    formula_cells = _formula_coordinates(formula_sheet, start_row, end_row, 1, max_col)

    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=max_col):
        values: dict[str, Any] = {}
        sources: dict[str, str] = {}
        for idx, cell in enumerate(row):
            header = normalized_headers[idx]
            values[header] = _serialize_cell_value(cell.value)
            sources[header] = cell.coordinate
        rows.append({
            "row_index": row[0].row if row else None,
            "values": values,
            "sources": sources,
        })

    warnings: list[str] = []
    if int(worksheet.max_column or 1) > _MAX_RETURN_COLUMNS:
        warnings.append(f"COLUMN_LIMIT_APPLIED: returned first {_MAX_RETURN_COLUMNS} columns")
    if formula_cells:
        warnings.append("FORMULA_CELLS_PRESENT: formulas were not executed; cached values were returned")
    truncated = (end_row < int(worksheet.max_row or 0)) or int(worksheet.max_column or 1) > _MAX_RETURN_COLUMNS

    return {
        "ok": True,
        "attachment_id": attachment_id,
        "operation": "extract_rows",
        "workbook": {
            "file_name": _attachment_name(attachment),
            "format": "xlsx",
            "sheet_count": len(workbook.sheetnames),
        },
        "sheet": {
            "name": sheet_name,
            "max_row": int(worksheet.max_row or 0),
            "max_column": int(worksheet.max_column or 0),
        },
        "header_row": header_row,
        "range": f"A{start_row}:{get_column_letter(max_col)}{end_row}",
        "columns": normalized_headers,
        "rows": rows,
        "truncated": truncated,
        "warnings": warnings,
    }


def _analyze_csv(
    content: bytes,
    attachment: dict[str, Any],
    attachment_id: str,
    operation: str,
    args: dict[str, Any],
) -> dict[str, Any]:
    rows = _csv_rows(content)
    max_row = len(rows)
    max_col = max((len(row) for row in rows), default=0)
    workbook_meta = {
        "file_name": _attachment_name(attachment),
        "format": "csv",
        "sheet_count": 1,
    }
    sheet_meta = {"name": "CSV", "max_row": max_row, "max_column": max_col}

    if operation == "summary":
        return {
            "ok": True,
            "attachment_id": attachment_id,
            "operation": "summary",
            "workbook": workbook_meta,
            "sheets": [{**sheet_meta, "state": "visible", "dimension": f"A1:{_column_letter(max_col)}{max_row}"}],
            "truncated": False,
            "warnings": [],
        }

    if operation == "preview":
        limit = _int_arg(args, "limit", _DEFAULT_PREVIEW_ROWS, 1, _MAX_RETURN_ROWS)
        return _csv_range_result(attachment_id, operation, workbook_meta, sheet_meta, rows, 1, 1, max_col, min(limit, max_row))

    if operation == "read_range":
        bounds = _parse_range(args.get("range") or args.get("cell_range") or args.get("cellRange"))
        if bounds is None:
            raise SpreadsheetAnalysisError("RANGE_REQUIRED", "valid range is required for read_range")
        min_col, min_row, max_col_range, max_row_range = bounds
        return _csv_range_result(
            attachment_id,
            operation,
            workbook_meta,
            sheet_meta,
            rows,
            min_row,
            min_col,
            max_col_range,
            max_row_range,
        )

    if operation == "extract_rows":
        header_row = _int_arg(args, "header_row", 1, 1, max(max_row, 1))
        offset = _int_arg(args, "offset", 0, 0, 100_000)
        limit = _int_arg(args, "limit", _DEFAULT_EXTRACT_ROWS, 1, _MAX_RETURN_ROWS)
        headers = _normalize_headers(_slice_csv_row(rows, header_row, 1, min(max_col, _MAX_RETURN_COLUMNS)))
        start_row = header_row + 1 + offset
        end_row = min(start_row + limit - 1, max_row)
        extracted: list[dict[str, Any]] = []
        for row_index in range(start_row, end_row + 1):
            row_values = _slice_csv_row(rows, row_index, 1, len(headers))
            values = {header: _truncate_value(row_values[idx] if idx < len(row_values) else "") for idx, header in enumerate(headers)}
            sources = {header: f"{_column_letter(idx + 1)}{row_index}" for idx, header in enumerate(headers)}
            extracted.append({"row_index": row_index, "values": values, "sources": sources})
        return {
            "ok": True,
            "attachment_id": attachment_id,
            "operation": "extract_rows",
            "workbook": workbook_meta,
            "sheet": sheet_meta,
            "header_row": header_row,
            "range": f"A{start_row}:{_column_letter(len(headers))}{end_row}",
            "columns": headers,
            "rows": extracted,
            "truncated": end_row < max_row or max_col > _MAX_RETURN_COLUMNS,
            "warnings": [] if max_col <= _MAX_RETURN_COLUMNS else [f"COLUMN_LIMIT_APPLIED: returned first {_MAX_RETURN_COLUMNS} columns"],
        }

    raise SpreadsheetAnalysisError("UNSUPPORTED_SPREADSHEET_OPERATION", operation)


def _csv_range_result(
    attachment_id: str,
    operation: str,
    workbook_meta: dict[str, Any],
    sheet_meta: dict[str, Any],
    rows: list[list[str]],
    min_row: int,
    min_col: int,
    max_col: int,
    max_row: int,
) -> dict[str, Any]:
    warnings: list[str] = []
    original_rows = max(0, max_row - min_row + 1)
    original_cols = max(0, max_col - min_col + 1)
    if original_rows > _MAX_RETURN_ROWS:
        max_row = min_row + _MAX_RETURN_ROWS - 1
        warnings.append(f"ROW_LIMIT_APPLIED: returned first {_MAX_RETURN_ROWS} rows from requested range")
    if original_cols > _MAX_RETURN_COLUMNS:
        max_col = min_col + _MAX_RETURN_COLUMNS - 1
        warnings.append(f"COLUMN_LIMIT_APPLIED: returned first {_MAX_RETURN_COLUMNS} columns from requested range")
    result_rows = []
    for row_index in range(min_row, min(max_row, len(rows)) + 1):
        values = _slice_csv_row(rows, row_index, min_col, max_col)
        result_rows.append({
            "row_index": row_index,
            "values": [_truncate_value(value) for value in values],
            "cells": [
                {
                    "coordinate": f"{_column_letter(min_col + idx)}{row_index}",
                    "value": _truncate_value(value),
                    "type": "string" if value != "" else "empty",
                }
                for idx, value in enumerate(values)
            ],
        })
    return {
        "ok": True,
        "attachment_id": attachment_id,
        "operation": operation,
        "workbook": workbook_meta,
        "sheet": sheet_meta,
        "range": f"{_column_letter(min_col)}{min_row}:{_column_letter(max_col)}{max_row}",
        "columns": [{"index": col, "letter": _column_letter(col)} for col in range(min_col, max_col + 1)],
        "rows": result_rows,
        "truncated": bool(original_rows > _MAX_RETURN_ROWS or original_cols > _MAX_RETURN_COLUMNS),
        "warnings": warnings,
    }


def _find_attachment(attachments: Any, attachment_id: str) -> dict[str, Any] | None:
    if not isinstance(attachments, list):
        return None
    for item in attachments:
        if not isinstance(item, dict):
            continue
        item_id = _text(item.get("id") or item.get("attachmentId") or item.get("attachment_id"))
        if item_id == attachment_id:
            return item
    return None


def _read_attachment_bytes(attachment: dict[str, Any], attachment_id: str) -> dict[str, Any]:
    inline = attachment.get("data") or attachment.get("content")
    if isinstance(inline, str) and inline.strip():
        try:
            raw = inline.split(",", 1)[1] if inline.strip().startswith("data:") and "," in inline else inline
            content = base64.b64decode(raw, validate=True)
        except Exception:
            return {"ok": False, "error": f"ATTACHMENT_INLINE_DATA_INVALID: {attachment_id}"}
        if len(content) > _MAX_ATTACHMENT_BYTES:
            return {"ok": False, "error": f"ATTACHMENT_TOO_LARGE: {attachment_id}"}
        return {"ok": True, "content": content}

    path = _text(
        attachment.get("path")
        or attachment.get("file_path")
        or attachment.get("filePath")
        or attachment.get("local_path")
        or attachment.get("localPath")
    )
    if path:
        try:
            file_path = Path(path)
            if file_path.stat().st_size > _MAX_ATTACHMENT_BYTES:
                return {"ok": False, "error": f"ATTACHMENT_TOO_LARGE: {attachment_id}"}
            return {"ok": True, "content": file_path.read_bytes()}
        except Exception as exc:
            logger.warning(
                "Spreadsheet local attachment read failed: attachment_id=%s error=%s",
                attachment_id,
                _safe_error_summary(exc),
            )
            return {"ok": False, "error": f"ATTACHMENT_READ_FAILED: {attachment_id}"}

    url = _text(
        attachment.get("url")
        or attachment.get("href")
        or attachment.get("downloadUrl")
        or attachment.get("download_url")
    )
    if not url:
        return {"ok": False, "error": f"ATTACHMENT_CONTENT_REQUIRED: {attachment_id}"}
    if not url.lower().startswith(("http://", "https://")):
        return {"ok": False, "error": f"UNSUPPORTED_ATTACHMENT_URL: {attachment_id}"}
    try:
        request = Request(url, headers={"User-Agent": "Foggy-BizWorker/1.0"})
        with urlopen(request, timeout=_DOWNLOAD_TIMEOUT_SECONDS) as response:
            length = response.headers.get("Content-Length")
            if length and int(length) > _MAX_ATTACHMENT_BYTES:
                return {"ok": False, "error": f"ATTACHMENT_TOO_LARGE: {attachment_id}"}
            content = response.read(_MAX_ATTACHMENT_BYTES + 1)
        if len(content) > _MAX_ATTACHMENT_BYTES:
            return {"ok": False, "error": f"ATTACHMENT_TOO_LARGE: {attachment_id}"}
        return {"ok": True, "content": content}
    except Exception as exc:
        logger.warning("Spreadsheet attachment download failed: attachment_id=%s error=%s", attachment_id, _safe_error_summary(exc))
        return {"ok": False, "error": f"ATTACHMENT_DOWNLOAD_FAILED: {attachment_id}"}


def _spreadsheet_format(attachment: dict[str, Any]) -> str | None:
    suffix = _attachment_suffix(attachment)
    mime_type = _text(
        attachment.get("mimeType")
        or attachment.get("mime_type")
        or attachment.get("contentType")
        or attachment.get("content_type")
    ).lower()
    if suffix == ".xlsx" or mime_type in _XLSX_MIME_TYPES:
        return "xlsx"
    if suffix == ".csv" or mime_type in _CSV_MIME_TYPES and suffix not in _UNSUPPORTED_EXCEL_SUFFIXES:
        return "csv"
    return None


def _attachment_suffix(attachment: dict[str, Any]) -> str:
    for value in (
        attachment.get("name"),
        attachment.get("fileName"),
        attachment.get("filename"),
        attachment.get("path"),
        attachment.get("file_path"),
        attachment.get("url"),
        attachment.get("href"),
        attachment.get("downloadUrl"),
        attachment.get("download_url"),
    ):
        text = _text(value)
        if not text:
            continue
        try:
            text = urlsplit(text).path or text
        except ValueError:
            text = text.split("?", 1)[0].split("#", 1)[0]
        suffix = Path(text).suffix.lower()
        if suffix:
            return suffix
    return ""


def _select_sheet(workbook: Any, requested_sheet: Any) -> Any:
    sheet_name = _text(requested_sheet)
    if not sheet_name:
        return workbook[workbook.sheetnames[0]]
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    for candidate in workbook.sheetnames:
        if candidate.lower() == sheet_name.lower():
            return workbook[candidate]
    raise SpreadsheetAnalysisError("SHEET_NOT_FOUND", sheet_name)


def _xlsx_row_values(worksheet: Any, row_index: int, min_col: int, max_col: int) -> list[Any]:
    for row in worksheet.iter_rows(min_row=row_index, max_row=row_index, min_col=min_col, max_col=max_col):
        return [_serialize_cell_value(cell.value) for cell in row]
    return []


def _formula_coordinates(
    formula_sheet: Any | None,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> dict[str, str]:
    if formula_sheet is None:
        return {}
    formulas: dict[str, str] = {}
    try:
        for row in formula_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas[cell.coordinate] = value
    except Exception:
        return formulas
    return formulas


def _cell_payload(cell: Any, *, formula_cells: dict[str, str], include_formula_text: bool) -> dict[str, Any]:
    payload = {
        "coordinate": cell.coordinate,
        "value": _serialize_cell_value(cell.value),
        "type": _cell_type(cell.value),
    }
    formula = formula_cells.get(cell.coordinate)
    if formula:
        payload["formula"] = True
        if include_formula_text:
            payload["formula_text"] = _truncate_value(formula)
    return payload


def _serialize_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, dt.time):
        return value.isoformat()
    if isinstance(value, str):
        return _truncate_value(value)
    if isinstance(value, (int, float, bool)):
        return value
    return _truncate_value(str(value))


def _cell_type(value: Any) -> str:
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, dt.datetime):
        return "datetime"
    if isinstance(value, dt.date):
        return "date"
    if isinstance(value, dt.time):
        return "time"
    return "string"


def _csv_rows(content: bytes) -> list[list[str]]:
    text = _decode_text(content)
    return list(csv.reader(io.StringIO(text)))


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _slice_csv_row(rows: list[list[str]], row_index: int, min_col: int, max_col: int) -> list[str]:
    if row_index < 1 or row_index > len(rows):
        return []
    row = rows[row_index - 1]
    result = []
    for index in range(min_col - 1, max_col):
        result.append(row[index] if index < len(row) else "")
    return result


def _parse_range(value: Any) -> tuple[int, int, int, int] | None:
    text = _text(value).upper()
    match = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", text)
    if not match:
        return None
    min_col = _column_index(match.group(1))
    min_row = int(match.group(2))
    max_col = _column_index(match.group(3))
    max_row = int(match.group(4))
    if min_col > max_col or min_row > max_row:
        return None
    return min_col, min_row, max_col, max_row


def _column_index(letters: str) -> int:
    value = 0
    for char in letters:
        value = value * 26 + ord(char) - ord("A") + 1
    return value


def _column_letter(index: int) -> str:
    if index <= 0:
        return "A"
    letters = ""
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _normalize_headers(headers: list[Any]) -> list[str]:
    result: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(headers, start=1):
        raw = _text(value)
        header = re.sub(r"\s+", "_", raw.strip()).strip("_") if raw else f"column_{index}"
        header = header or f"column_{index}"
        base = header
        count = seen.get(base, 0) + 1
        seen[base] = count
        if count > 1:
            header = f"{base}_{count}"
        result.append(header)
    return result


def _int_arg(args: dict[str, Any], name: str, default: int, minimum: int, maximum: int) -> int:
    value = args.get(name)
    camel = "".join([name.split("_")[0], *[part.capitalize() for part in name.split("_")[1:]]])
    if value is None:
        value = args.get(camel)
    try:
        number = int(value) if value is not None else default
    except (TypeError, ValueError):
        number = default
    return min(max(number, minimum), maximum)


def _bool_option(args: dict[str, Any], name: str, default: bool) -> bool:
    options = args.get("options")
    value = options.get(name) if isinstance(options, dict) and name in options else args.get(name)
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return _text(value).lower() in {"true", "1", "yes", "y"}


def _maybe_externalize_result(
    result: dict[str, Any],
    *,
    artifact_store: Any | None,
    account_id: str | None,
    task_id: str | None,
) -> dict[str, Any]:
    encoded = json.dumps(result, ensure_ascii=False)
    if len(encoded) <= _MAX_RESULT_CHARS:
        return result
    compact = dict(result)
    compact["rows"] = result.get("rows", [])[:5] if isinstance(result.get("rows"), list) else result.get("rows")
    compact["truncated"] = True
    warnings = list(result.get("warnings") or [])
    warnings.append("RESULT_SIZE_LIMIT_APPLIED: full result was externalized or compacted")
    compact["warnings"] = warnings
    if artifact_store is None or not account_id:
        return compact
    try:
        artifact = artifact_store.create(
            account_id=account_id,
            task_id=task_id,
            scope="task",
            name=f"{result.get('attachment_id', 'spreadsheet')}-{result.get('operation', 'result')}.json",
            content=json.dumps(result, ensure_ascii=False, indent=2),
            mime_type="application/json",
            encoding="utf-8",
            summary="Full spreadsheet analysis result externalized due to tool result size limit.",
        )
        compact["artifact"] = artifact
    except Exception as exc:
        warnings.append(f"ARTIFACT_WRITE_FAILED: {_safe_error_summary(exc)}")
        compact["warnings"] = warnings
    return compact


def _attachment_name(attachment: dict[str, Any]) -> str:
    name = _text(attachment.get("name") or attachment.get("fileName") or attachment.get("filename"))
    return name or "spreadsheet"


def _truncate_value(value: Any) -> str:
    text = "" if value is None else str(value)
    return text if len(text) <= _MAX_CELL_CHARS else text[:_MAX_CELL_CHARS] + "...[truncated]"


def _safe_error_summary(exc: Exception) -> str:
    text = str(exc).splitlines()[0].strip() or exc.__class__.__name__
    text = re.sub(r"https?://[^\s'\"<>]+", "[redacted-url]", text)
    text = re.sub(
        r"(?i)\b(api[_-]?key|apikey|token|secret|signature|credential|password)\s*[:=]\s*['\"]?[^'\"\s,;]+",
        r"\1=[redacted]",
        text,
    )
    text = re.sub(r"(?i)(sk|ak)-[A-Za-z0-9_\-]{8,}", "[redacted-key]", text)
    return text[:240]


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
